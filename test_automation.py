from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from pathlib import Path
import argparse
import os
import re
import sys

ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_URL = "https://www.pixelssuite.com/chat-translator"


def configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass


def resolve_path(path_text: str) -> str:
    p = Path(path_text)
    if p.is_absolute():
        return str(p)
    return str((ROOT_DIR / p).resolve())


def normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_header_row(ws) -> int:
    """Find the row that contains Input and Expected output headers."""
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [normalize_header(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)]
        if "input" in values and "expectedoutput" in values:
            return row
    return 1


def find_col(headers, possible_names):
    normalized = [normalize_header(v) for v in headers]
    for name in possible_names:
        key = normalize_header(name)
        if key in normalized:
            return normalized.index(key) + 1
    return None


def get_real_cell(ws, row, col):
    """If a cell is merged, return the top-left writable cell."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def cell_text(ws, row, col) -> str:
    value = get_real_cell(ws, row, col).value
    return "" if value is None else str(value).strip()


def set_cell(ws, row, col, value):
    get_real_cell(ws, row, col).value = value


def read_output(output_box) -> str:
    try:
        value = output_box.input_value(timeout=3000)
        return value.strip() if value else ""
    except Exception:
        try:
            return output_box.inner_text(timeout=3000).strip()
        except Exception:
            return ""


def clear_and_type(page, input_box, text: str, type_delay_ms: int):
    input_box.click()
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    input_box.type(text, delay=max(0, int(type_delay_ms)))


def wait_for_new_output(page, previous_output: str, timeout_ms: int):
    """Wait until the second textarea has a non-empty output different from previous output."""
    try:
        page.wait_for_function(
            """
            (previous) => {
                const textareas = Array.from(document.querySelectorAll('textarea'));
                const output = (textareas[1]?.value || '').trim();
                const buttons = Array.from(document.querySelectorAll('button'));
                const busy = buttons.some(b => (b.innerText || '').toLowerCase().includes('transliterating'));
                return output.length > 0 && output !== previous && !busy;
            }
            """,
            arg=previous_output,
            timeout=timeout_ms,
        )
        return True
    except PlaywrightTimeoutError:
        return False


def parse_args():
    parser = argparse.ArgumentParser(description="Run Singlish to Sinhala transliteration test cases from Excel.")
    parser.add_argument("--excel", required=True, help="Excel file path")
    parser.add_argument("--sheet", default="Test cases", help="Worksheet name. Defaults to 'Test cases'.")
    parser.add_argument("--url", default=DEFAULT_URL, help="Translator URL")
    parser.add_argument("--wait-ms", type=int, default=5000, help="Extra wait after output appears")
    parser.add_argument("--type-delay-ms", type=int, default=80, help="Typing delay per character")
    parser.add_argument("--slow-mo-ms", type=int, default=200, help="Playwright slow motion delay")
    parser.add_argument("--output-timeout-ms", type=int, default=30000, help="Output wait timeout")
    parser.add_argument("--save-every", type=int, default=1, help="Save Excel after this many rows")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    parser.add_argument("--keep-open", action="store_true", help="Keep browser open at the end")
    return parser.parse_args()


def main():
    configure_stdout()
    args = parse_args()
    excel_path = resolve_path(args.excel)

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        return

    try:
        workbook = load_workbook(excel_path)
    except PermissionError:
        print("ERROR: Close the Excel file before running the script, then try again.")
        return
    except Exception as exc:
        print(f"ERROR: Could not open Excel file: {exc}")
        return

    ws = workbook[args.sheet] if args.sheet in workbook.sheetnames else workbook.active
    header_row = find_header_row(ws)
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

    input_col = find_col(headers, ["Input"])
    expected_col = find_col(headers, ["Expected output", "Expected Output"])
    actual_col = find_col(headers, ["Actual output", "Actual Output"])
    status_col = find_col(headers, ["Status"])

    if not input_col or not expected_col:
        print("ERROR: Excel must contain 'Input' and 'Expected output' columns.")
        return

    if not actual_col:
        actual_col = ws.max_column + 1
        ws.cell(row=header_row, column=actual_col).value = "Actual output"
    if not status_col:
        status_col = ws.max_column + 1
        ws.cell(row=header_row, column=status_col).value = "Status"

    rows = [r for r in range(header_row + 1, ws.max_row + 1) if cell_text(ws, r, input_col)]
    print(f"Found {len(rows)} test cases in: {Path(excel_path).name}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=max(0, args.slow_mo_ms))
        page = browser.new_page()
        page.set_default_timeout(60000)

        print(f"Opening site: {args.url}")
        page.goto(args.url, wait_until="domcontentloaded")
        page.wait_for_selector("textarea", timeout=60000)

        input_box = page.locator("textarea").nth(0)
        output_box = page.locator("textarea").nth(1)
        transliterate_button = page.get_by_role("button", name=re.compile(r"Transliterate", re.I))

        processed = 0
        for row in rows:
            tc_id = cell_text(ws, row, 1) or f"Row {row}"
            singlish_input = cell_text(ws, row, input_col)
            expected_output = cell_text(ws, row, expected_col)

            print(f"\n[{tc_id}] Input: {singlish_input}")
            try:
                previous_output = read_output(output_box)
                clear_and_type(page, input_box, singlish_input, args.type_delay_ms)
                transliterate_button.click()

                output_ready = wait_for_new_output(page, previous_output, args.output_timeout_ms)
                page.wait_for_timeout(max(0, args.wait_ms))

                actual_output = read_output(output_box)
                status = "PASS" if actual_output == expected_output else "FAIL"
                if not output_ready:
                    print("  Warning: Output may not have fully refreshed before timeout.")

                set_cell(ws, row, actual_col, actual_output)
                set_cell(ws, row, status_col, status)

                print(f"  Actual: {actual_output}")
                print(f"  Status: {status}")

            except Exception as exc:
                set_cell(ws, row, status_col, "UI ERROR")
                print(f"  ERROR: {exc}")

            processed += 1
            if args.save_every > 0 and processed % args.save_every == 0:
                workbook.save(excel_path)

        workbook.save(excel_path)
        print("\nCompleted. Results saved to Excel.")

        if args.keep_open and not args.headless:
            print("Browser is kept open. Press Ctrl+C in terminal when finished.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                pass

        browser.close()


if __name__ == "__main__":
    main()
